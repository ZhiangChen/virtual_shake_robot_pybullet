import os
import yaml
import pandas as pd
from openpyxl import Workbook, load_workbook
from launch.substitutions import LaunchConfiguration, PathJoinSubstitution
from launch.actions import DeclareLaunchArgument, GroupAction
from launch import LaunchDescription
from launch_ros.actions import Node, PushRosNamespace
from launch_ros.substitutions import FindPackageShare

def calculate_objective_function(accuracy, precision, recall, weights=(0.33, 0.33, 0.33)):
    """
    Calculate the objective function as a weighted combination of accuracy, precision, and recall.
    """
    w1, w2, w3 = weights
    return w1 * accuracy + w2 * precision + w3 * recall

def deep_sanitize(data):
    """
    Recursively sanitize data to ensure all elements are YAML serializable.
    """
    if isinstance(data, dict):
        return {deep_sanitize(key): deep_sanitize(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [deep_sanitize(item) for item in data]
    elif isinstance(data, (int, float, str)):
        return data
    else:
        return str(data)  # Convert unsupported types to string

def generate_yaml_files(output_dir, base_content, parameter_combinations, model_choice):
    """
    Generate YAML files for given parameter combinations, including the mesh file path.
    """
    os.makedirs(output_dir, exist_ok=True)

    # Set the correct mesh file path based on the model choice
    ros2_ws = os.getenv('ROS2_WS', default=os.path.expanduser('~/ros2_ws'))
    if model_choice == 'sp1':
        mesh_path = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet/models/SP1_PBRmodel/sp1.urdf')
    elif model_choice == 'sp2':
        mesh_path = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet/models/SP2_PBRmodel/sp2.urdf')
    else:  # Default to pbr_mesh
        mesh_path = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet/models/double_rock_pbr/pbr_mesh.urdf')

    yaml_files = []

    for idx, params in enumerate(parameter_combinations):
        # Start with the base content and update the parameters
        content = base_content.copy()
        content['/**']['ros__parameters']['rock_structure_mesh'].update(params)
        content['/**']['ros__parameters']['rock_structure_mesh']['mesh'] = mesh_path  # Include the mesh path

        # Name each YAML file sequentially
        filename = f"{model_choice}_{idx}.yaml"
        file_path = os.path.join(output_dir, filename)

        # Write the YAML file
        with open(file_path, 'w') as yaml_file:
            yaml.safe_dump(content, yaml_file, default_flow_style=None, sort_keys=False, indent=2, width=120)

        print(f"Generated YAML: {file_path}")
        yaml_files.append(file_path)

    return yaml_files


def generate_launch_files(launch_dir, yaml_files, model_choice):
    """
    Generate ROS2 launch files dynamically for the given YAML files and model choice.
    """
    ros2_ws = os.getenv('ROS2_WS', default=os.path.expanduser('~/ros2_ws'))
    config_directory = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet/config')  # Replace {{ROS2_WS}}
    launch_file_paths = []

    for i, yaml_file in enumerate(yaml_files):
        # Calculate dynamic namespace
        sim_ns = f"sim_{i+1}"

        # Generate the launch file content
        launch_content = f"""
#!/usr/bin/env python3
from launch.substitutions import LaunchConfiguration
from launch.actions import DeclareLaunchArgument, GroupAction
from launch import LaunchDescription
from launch_ros.actions import Node, PushRosNamespace
import os

def generate_launch_description():

    # Declare the motion_mode argument
    motion_mode_arg = DeclareLaunchArgument(
        'motion_mode',
        default_value='',
        description='Motion mode for the control node: grid_cosine, single_recording, all_recordings'
    )

    # Paths to configuration files
    physics_engine_parameters_path = os.path.join('{config_directory}', 'physics_engine_parameters.yaml')
    physics_parameters_path = os.path.join('{config_directory}', 'physics_parameters.yaml')
    vsr_structure_path = os.path.join('{config_directory}', 'vsr_structure_box.yaml')
    pbr_structure_path = os.path.join('{config_directory}', 'pbr_box.yaml')

    # Perception node for namespace {sim_ns}
    perception_node = Node(
        package='virtual_shake_robot_pybullet',
        executable='perception_node.py',
        name='perception_node',
        namespace='{sim_ns}',
        output='screen'
    )

    # Simulation node for namespace {sim_ns} using the YAML file
    simulation_node = Node(
        package='virtual_shake_robot_pybullet',
        executable='simulation_node.py',
        name='simulation_node',
        namespace='{sim_ns}',
        output='screen',
        parameters=[
            physics_engine_parameters_path,
            physics_parameters_path,
            vsr_structure_path,
            pbr_structure_path,
            '{yaml_file}',
            {{'recording_folder_name': LaunchConfiguration('recording_folder_name')}}
        ]
    )

    # Control node for namespace {sim_ns}
    control_node = Node(
        package='virtual_shake_robot_pybullet',
        executable='control_node.py',
        name='control_node',
        namespace='{sim_ns}',
        output='screen',
        parameters=[
            {{'motion_mode': LaunchConfiguration('motion_mode')}}
        ]
    )

    return LaunchDescription([
        motion_mode_arg,
        GroupAction([
            PushRosNamespace('{sim_ns}'),
            perception_node,
            simulation_node,
            control_node    
        ])
    ])
        """

        # Save the launch file in the specified launch_dir
        launch_filename = os.path.join(launch_dir, f"{model_choice}_{i}.launch.py")
        with open(launch_filename, 'w') as file:
            file.write(launch_content)

        launch_file_paths.append(launch_filename)
        print(f"Generated: {launch_filename}")

    return launch_file_paths



def generate_excel_file(output_file, data, iteration, parameter_name, fixed_params, overwrite=True):
    """
    Generate or update an Excel file to log parameters and results for each iteration.
    
    Parameters:
    - output_file: Path to the Excel file.
    - data: Data from the current iteration.
    - iteration: Current iteration number.
    - parameter_name: The parameter being optimized in this iteration.
    - fixed_params: Dictionary of fixed parameter values.
    - overwrite: If True, overwrite the existing file.
    """
    # Define the required column order
    ordered_columns = [
        "Iteration", "restitution", "lateralFriction", "spinningFriction", 
        "contactDamping", "contactStiffness", "Accuracy", "Precision", "Recall", "Objective"
    ]
    
    # Extract only the required parameters and maintain order
    param_values = {k: fixed_params.get(k, "N/A") for k in ordered_columns[1:-4]}  # Exclude Iteration & Metrics

    if overwrite or not os.path.exists(output_file):
        # Create a new Excel file or overwrite the existing one
        wb = Workbook()
        ws = wb.active
        ws.title = "Optimization Log"

        # Add headers in the defined order
        for col, header in enumerate(ordered_columns, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(output_file)

    # Load the existing Excel file
    wb = load_workbook(output_file)
    ws = wb["Optimization Log"]

    # Add data rows
    for row in data:
        full_row = [
            iteration,  # Current iteration number
            param_values["restitution"],  
            param_values["lateralFriction"],  
            param_values["spinningFriction"],  
            param_values["contactDamping"],  
            param_values["contactStiffness"],  
            row["Accuracy"],  
            row["Precision"],  
            row["Recall"],  
            row["Objective"]
        ]
        ws.append(full_row)

    # Save the updated Excel file
    wb.save(output_file)
    print(f"Updated Excel file: {output_file}")





def iterative_optimization_with_excel(
    csv_file, base_params, parameter_ranges, model_choice, output_dir, launch_dir, excel_file
):
    """
    Perform iterative optimization by testing each parameter range sequentially,
    and record results in an Excel file.
    """
    fixed_params = base_params.copy()
    iteration_results = []  # To store results of all iterations

    for iteration, (param, values) in enumerate(parameter_ranges.items(), start=1):
        print(f"\nOptimizing parameter: {param} over range {values}")

        # Generate parameter combinations for this iteration
        parameter_combinations = []
        for value in values:
            current_params = fixed_params.copy()
            current_params[param] = value  # Override only the current parameter
            parameter_combinations.append(current_params)

        # Generate YAML files for this parameter
        base_content = {
            '/**': {
                'ros__parameters': {
                    'rock_structure_mesh': {}
                }
            }
        }
        yaml_files = generate_yaml_files(output_dir, base_content, parameter_combinations, model_choice)

        generate_launch_files(launch_dir=launch_dir, yaml_files=yaml_files, model_choice=model_choice)

        # Simulate or wait for updated CSV with results for this parameter
        print("\nRun experiments using the generated files. Update the CSV file with accuracy, precision, and recall.")
        input("Press Enter after updating the CSV file to continue optimization...")

        # Load updated results
        data = pd.read_csv(csv_file)
        data.columns = [col.strip().replace(" ", "") for col in data.columns]  # Remove spaces in column names

        # Filter the DataFrame for the current parameter
        data_filtered = data[data['Parameter'] == param]

        # Calculate the objective function for the filtered data
        data_filtered['objective'] = calculate_objective_function(
            data_filtered['Accuracy'], 
            data_filtered['Precision'], 
            data_filtered['Recall']
        )

        # Append results for each row to iteration_results
        for _, row in data_filtered.iterrows():
            iteration_results.append({
                'Value': row['Value'],  # Use the 'Value' column for the parameter value
                'Accuracy': row['Accuracy'],
                'Precision': row['Precision'],
                'Recall': row['Recall'],
                'Objective': row['objective']
            })

        # Find the best value for the current parameter
        best_row = data_filtered.loc[data_filtered['objective'].idxmax()]
        best_value = best_row['Value']
        print(f"Best value for {param}: {best_value} with objective score: {best_row['objective']}")

        # Fix the best value for this parameter
        fixed_params[param] = float(best_value)

        # Record results in the Excel file, including all parameters
        generate_excel_file(
            output_file=excel_file,
            data=iteration_results,
            iteration=iteration,
            parameter_name=param,
            fixed_params=fixed_params
        )

    print(f"\nFinal optimized parameters: {fixed_params}")
    return fixed_params


def main_with_excel():
    # Define initial base parameters
    base_params = {
        'meshScale': [1.0, 1.0, 1.0],
        'mass': 353.802,
        'restitution': 0.3,
        'lateralFriction': 0.5,
        'spinningFriction': 0.1,
        'contactDamping': 10000.0,
        'contactStiffness': 1000000.0,
        'rock_position': [0.0, 0.0, 2.78]
    }

    # Define parameter ranges
    parameter_ranges = {
        'restitution': [0.3, 0.35, 0.40, 0.45, 0.5, 0.6, 0.7, 0.8, 0.9],
        'lateralFriction': [0.31, 0.41, 0.51, 0.55, 0.66, 0.7, 0.8, 1.0, 1.5],
        'spinningFriction': [0.01, 0.05, 0.1, 0.2],
        'contactDamping': [100000.0, 500000.0, 1000000.0],
        'contactStiffness': [1000000.0, 5000000.0, 7000000.0]
    }

    # Model choice
    model_choice = 'sp1'  # Modify based on your model
    ros2_ws = os.getenv('ROS2_WS', default=os.path.expanduser('~/ros2_ws'))
    output_directory = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet', 'data')
    launch_directory = os.path.join(ros2_ws, 'src', 'virtual_shake_robot_pybullet', 'launch')
    csv_file_path = os.path.join(output_directory, 'simulation_metrics.csv')  # Path to the CSV file
    excel_file_path = os.path.join(output_directory, 'optimization_results.xlsx')  # Path to the Excel file

    # Perform iterative optimization and record results in Excel
    iterative_optimization_with_excel(
        csv_file=csv_file_path,
        base_params=base_params,
        parameter_ranges=parameter_ranges,
        model_choice=model_choice,
        output_dir=output_directory,
        launch_dir=launch_directory,
        excel_file=excel_file_path
    )

if __name__ == "__main__":
    main_with_excel()
